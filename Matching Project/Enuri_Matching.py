import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import json
import xlsxwriter
from googletrans import Translator

#===================
#
#Connector for enuri
#
#===================

proxy_support = {'http': 'http://11115:7My2Ng@world.nohodo.com:6811'}
#
# session = requests.Session()
# session.headers = ({
#             'Host': 'www.danawa.com',
#             'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36',
#             'Upgrade-Insecure-Requests': '1',
#             'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
#             'Accept-Encoding': 'gzip, deflate',
#             'Accept-Language': 'en-GB,en;q=0.9'})
#====================code to reac excel===================


SplChar = {'!': '%21', '#': '%23', '$': '%24', '&': '%26', "'": '%27', '(': '%28', ')': '%29', '*': '%2A', '+': '%2B', ',': '%2C', '/': '%2F', ':' : '%3A', ';': '%3B', '=': '%3D',  '?': '%3F',  '@': '%40',  '[': '%5B',  ']': '%5D',  '"': '%22',  '%': '%25',  '<': '%3C',  '>': '%3E'}
dataList = []

wb = load_workbook(filename='E:\\Hemant Python\\Python Test\\Danawa_Data\\matcing_itemnumber.xlsx')
ws = wb['Sheet1']
row = ws.max_row
col = ws.max_column
for row in ws.iter_rows(min_row=2, min_col=0, max_row=row, max_col=col):
    data = []
    for cell in row:
        data.append(cell.value)
        # print(data)
    dataList.append(data)
wb.close()

#=================homepage hit
# Homepage ="http://www.danawa.com/"
# homeRes = requests.get(Homepage, proxies=proxy_support, headers=session.headers)


#=============Data Out Excel
# workbook = xlsxwriter.Workbook('E:\\Software\\hemant\\Python Test\\Danawa_Data\\TestData.xlsx')
# worksheet = workbook.add_worksheet()
# wrk_row = 0
# wrk_col = 0
# worksheet.write(wrk_row, wrk_col, 'Brand')
# worksheet.write(wrk_row, wrk_col+1, 'Product Name')
# worksheet.write(wrk_row, wrk_col+2, 'Dell online price')
# worksheet.write(wrk_row, wrk_col+3, 'Search URL')
# worksheet.write(wrk_row, wrk_col+4, 'URL')
# worksheet.write(wrk_row, wrk_col+5, 'Name')
# worksheet.write(wrk_row, wrk_col+6, 'Translated Name')
# worksheet.write(wrk_row, wrk_col+7, 'Price')
# worksheet.write(wrk_row, wrk_col+8, 'Discount Price')
# worksheet.write(wrk_row, wrk_col+9, 'Price Difference')
# worksheet.write(wrk_row, wrk_col+10, 'Status')
# wrk_row += 1
#
# workbook.close()
#============================

for d in dataList:
    finalData = []
    minIndex = 0
    minprice = 0
    index = 1
    Brand = str(d[0])
    ProductName = str(d[1])
    Dprice = int(d[2])
    text = ProductName
    for c in text:
        if c in SplChar:
            text = text.replace(c, SplChar[c])
    temp = ProductName.replace(" ", '+')
    #=======================
    trans = Translator()
    #=======================
    SearchUrl = "http://www.enuri.com/lsv2016/ajax/getSearchGoods_ajax.jsp?random_seq=394&pageNum=1&pageGap=30&tabType=0&cate=&IsDeliverySumPrice=N&keyword="+text+"&in_keyword=&IsJungoPriceRemove=N&m_price=&start_price=&end_price="
    searchRes = requests.get(SearchUrl, proxies=proxy_support)
    soup = BeautifulSoup(searchRes.text, "lxml")
    json_data = json.loads(searchRes.text)
    # print(json_data.keys())
    if 'srpModelList' in json_data:
        ProdDiv = json_data['srpModelList']
        # print(type(ProdDiv))       0........
        for pr in ProdDiv:
            PrData = {'Index': '',
                      'Brand': '',
                      'ProductName': '',
                      'Dprice': '',
                      'SearchURL': '',
                      'URL': '',
                      'pr_name': '',
                      'tr_name': '',
                      'ListPrice': '',
                      'Disc_Price': '',
                      'Difference': '',
                      'CheckCount': '',
                      'Status': ''}
            count = 0
            # print(pr)
            URL = "http://www.enuri.com/detail.jsp?modelno=" + pr['intModelNo']
            pr_name = str(pr['strModelName'])
            # tr_text = str(trans.translate(pr_name, src='ko'))
            strt = tr_text.index('text')
            end = tr_text.index(', pronunciation')
            tr_text = tr_text[strt + 5:end]
            pr_Brand = pr['strBrand']
            # if Brand.lower() in tr_text.lower():
            #     textSplit = text.split(' ')
            #     for t in textSplit:
            #         if t in tr_text:
            #             count += 1

            priceTxt = pr['lngMinPrice'].replace(".", "")
            Price = int(priceTxt)
            # if count >= 2:
            #     if Dprice > Price:
            #         DiffPrice = Dprice - Price
            #     else:
            #         DiffPrice = Price - Dprice
            #     checkText = 'Valid'
            # else:
            #     checkText = 'Invalid'
            # if minprice == 0 and minIndex == 0:
            #     minIndex = index
            #     minprice = int(DiffPrice)
            PrData['Index'] = index
            PrData['Brand'] = Brand
            PrData['ProductName'] = ProductName
            PrData['Dprice'] = Dprice
            PrData['SearchUrl'] = SearchUrl
            PrData['URL'] = URL
            PrData['URL'] = pr_name
            # PrData['tr_text'] = tr_text
            PrData['ListPrice'] = Price
            PrData['Disc_Price'] = Price
            # PrData['Difference'] = DiffPrice
            # PrData['CheckCount'] = count
            PrData['Status'] = ''
            finalData.append(PrData)
            print(tr_text)
        #find min in final data for single search
        # for fd in finalData:
        #     # diffPrice = fd['Difference']
        #     if int(fd['Difference']) < minprice and int(fd['CheckCount']) >= 2:
        #         temp = finalData[minIndex]
        #         temp['Status'] = 'Invalid'
        #         minprice = fd['Difference']
        #         minIndex = fd['Index']
        #     elif int(fd['Difference']) == minprice and int(fd['CheckCount']) >= 2:
        #         temp = finalData[minIndex]
        #         temp['Status'] = 'Invalid'
        #         minprice = fd['Difference']
        #         minIndex = fd['Index']
        #     else:
        #         fd['Status'] = 'Invalid'




print("thank you")

