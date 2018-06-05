import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import xlsxwriter

#===================
#
#Connector for enuri
#
#===================
Dataout=[]
proxy_support = {'http': 'http://11115:7My2Ng@world.nohodo.com:6811'}
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/64.0.3282.186 Safari/537.36',
           'content-type': 'text/html;charset=UTF-8',
           'accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
           'Host': 'www.amazon.in'}
SplChar = {'!': '%21', '#': '%23', '$': '%24', '&': '%26', "'": '%27', '(': '%28', ')': '%29', '*': '%2A', '+': '%2B', ',': '%2C', '/': '%2F', ':' : '%3A', ';': '%3B', '=': '%3D',  '?': '%3F',  '@': '%40',  '[': '%5B',  ']': '%5D',  '"': '%22',  '%': '%25',  '<': '%3C',  '>': '%3E'}
dataList = []
wb = load_workbook(filename='E:\Hemant Python\MatchingConnectors\Amazon.in\Amazon.in_inputs.xlsx')
ws = wb['Sheet1']
row = ws.max_row
col = ws.max_column
for row in ws.iter_rows(min_row=2, min_col=0, max_row=row, max_col=col):
    for cell in row:
        dataList.append(cell.value)
wb.close()

workbook = xlsxwriter.Workbook(filename='E:\Hemant Python\MatchingConnectors\Amazon.in\Amazon.in_Dataout.xlsx')
worksheet = workbook.add_worksheet('DataOut')
wrk_row = 0
wrk_col = 0

worksheet.write(wrk_row, wrk_col, 'Product Name')
worksheet.write(wrk_row, wrk_col+1, 'Search URL')
worksheet.write(wrk_row, wrk_col+2, 'URL')
worksheet.write(wrk_row, wrk_col+3, 'Name')
worksheet.write(wrk_row, wrk_col+4, 'Original Price')
worksheet.write(wrk_row, wrk_col+5, 'Discount Price')
wrk_row += 1


for item in dataList:
    temp = item.strip().replace(' ', '+')
    url = 'https://www.amazon.in/s/ref=nb_sb_noss_2?url=search-alias%3Daps&field-keywords=' + temp + '&rh=i%3Aaps%2Ck%3A'+temp
    print(url)
    # Make the initial request and get the first response.
    r = requests.get(url, proxies = proxy_support, headers= headers).text
    # html = html.fromstring(r)
    # #print(r)
    # allProducts=html.xpath("//li[@class='s-result-item celwidget  ']")
    # for prod in allProducts

    soup = BeautifulSoup(r, "lxml")
    if "We found 0 results" in str(r):
        continue
    ProdUL = soup.find("ul", {"id": "s-results-list-atf"})
    ProdClass = ProdUL.find_all("li", {"class": "s-result-item celwidget "})
    for prod in ProdClass:
        data = []
        data.append(item)
        data.append(url)
        Url = prod.find("a", {"class": "a-link-normal s-access-detail-page s-color-twister-title-link a-text-normal"})
        ProductUrl = Url['href']
        data.append(ProductUrl)
        print(ProductUrl)
        ProdName = prod.find("h2").text.strip()
        data.append(ProdName)
        OriginalPrice = prod.find("span", {"class": "a-size-small a-color-secondary a-text-strike"})
        if OriginalPrice is None:
            OP = ""
        else:
            OP = OriginalPrice.text
        data.append(OP.replace(",","").replace(" ",""))
        DiscountedPrice = prod.find("span", {"class": "a-size-base a-color-price s-price a-text-bold"})
        if DiscountedPrice is None:
            DP = ""
        else:
            DP = DiscountedPrice.text
        data.append(DP.replace(",", "").replace(" ", ""))
        Dataout.append(data)
    for out in Dataout:
        worksheet.write(wrk_row, wrk_col, out[0])
        worksheet.write(wrk_row, wrk_col + 1, out[1])
        worksheet.write(wrk_row, wrk_col + 2, out[2])
        worksheet.write(wrk_row, wrk_col + 3, out[3])
        worksheet.write(wrk_row, wrk_col + 4, out[4])
        worksheet.write(wrk_row, wrk_col + 5, out[5])
        wrk_row += 1
    wrk_row += 2

workbook.close()