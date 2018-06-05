# -*- coding: utf-8 -*-
"""
Created on Wed Jan 10 13:06:46 2018

@author: Tech
"""
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

UserAgent = "Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/63.0.3239.132 Safari/537.36"
token=""
CatNo=[]
CatName=[]

proxy_support = {'https': 'https://csimonra:h19VA2xZ@173.246.181.89:29842',
                 'https': 'https://csimonra:h19VA2xZ@208.110.59.137:29842',
                 'https': 'https://csimonra:h19VA2xZ@208.110.59.208:29842',
                 'https': 'https://csimonra:h19VA2xZ@217.182.223.112:29842',
                 'https': 'https://csimonra:h19VA2xZ@217.182.223.118:29842',
                 'https': 'https://csimonra:h19VA2xZ@217.182.223.126:29842',
                 'https': 'https://csimonra:h19VA2xZ@173.228.169.148:29842'}
#==============================================================================
wb = load_workbook(filename = 'E:\\Souk\\Souk_Category.xlsx')
ws = wb['Sheet1']
row = ws.max_row
col = ws.max_column
for row in ws.iter_rows(min_row=1, min_col=0, max_row=row, max_col=1):
    for cell in row:
        CatNo.append(cell.value)
row = ws.max_row
for row1 in ws.iter_rows(min_row=1, min_col=2, max_row=row, max_col=2):
    for cell in row1:
        CatName.append(cell.value)
#==============================================================================

with requests.session() as c:
    homeurl = "https://sell.souq.com/sell"
    response = c.get(homeurl, proxies=proxy_support)
    token = c.cookies['SCXAT']
    token = token[:token.find('+')]
    print(token)
    loginurl = "https://sell.souq.com/login/validate"
    logindata = "params%5Bemail%5D=jtrinca%40souq.com&params%5Bpassword%5D=eclerx123&params%5Bremember_me%5D=off&token="+token
    headers = {'Accept':'Accept: application/json, text/plain, */*','Content-Type':'application/x-www-form-urlencoded;charset=UTF-8','Referer':'https://sell.souq.com/sell','Origin':'https://sell.souq.com','Host':'sell.souq.com','User-Agent':UserAgent}
    response1 = c.post(loginurl,data = logindata,headers = headers ,proxies =proxy_support)
    dash = open("E:\\Souk\\dash.html","w+")
    dash.write(response1.text)
    dash.close()
    dwnldURL = "https://sell.souq.com/inventory/downloadTemplate"
    postdata = "data%5BcreateNewItemDetails%5D%5B0%5D%5BitemTypeId%5D="+"3"+"&token="+token
    print(postdata)
    headers = {'Accept':'*/*','Content-Type':'application/x-www-form-urlencoded; charset=UTF-8','Referer':'https://sell.souq.com/inventory/search-item','Origin':'https://sell.souq.com','Host':'sell.souq.com','User-Agent':UserAgent}
    response2 = requests.post(dwnldURL,data = postdata,headers = headers,cookies=Homecookie,proxies =proxy_support)
    print(response2.text)
