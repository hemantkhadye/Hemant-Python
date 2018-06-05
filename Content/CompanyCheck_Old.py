# -*- coding: utf-8 -*-
"""
Created on Fri Jan 12 13:02:24 2018

@author: Tech
"""

import requests
import os
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import xlsxwriter
#import random

UserAgent = "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.94 Safari/537.36"
Company = "Global Payments U.K. Ltd"
compname = Company.replace(" ", "+")
urlname = []
#==============================================================================
# =============================================================================
# proxy=[]
# port='29842'
# uname='csimonra'
# password='h19VA2xZ'
# wb= load_workbook(filename = 'E:\\Company\\Input\\11Jan.xlsx')
# ws = wb['Sheet1']
# row = ws.max_row
# col=ws.max_column
# for row in ws.iter_rows(min_row=2, min_col=0, max_row=row, max_col=1):
#     for cell in row:            
#         proxy.append(cell.value)
# def getproxy():
#     rno = random.randint(0, row)
#     prx = proxy[rno]
#     proxy_support ={'https': 'https://'+uname+':'+password+'@'+prx+':'+port+'}
#     return proxy_support
# =============================================================================
proxy_support = {'https':'https://csimonra:h19VA2xZ@173.246.181.89:29842',
                 'https':'https://csimonra:h19VA2xZ@208.110.59.137:29842',
                 'https':'https://csimonra:h19VA2xZ@208.110.59.208:29842',
                 'https':'https://csimonra:h19VA2xZ@217.182.223.112:29842',
                 'https':'https://csimonra:h19VA2xZ@217.182.223.118:29842',
                 'https':'https://csimonra:h19VA2xZ@217.182.223.126:29842',
                 'https':'https://csimonra:h19VA2xZ@173.228.169.148:29842'}
#==============================================================================
homeurl = "https://companycheck.co.uk/"
response = requests.get(homeurl,proxies = proxy_support)
Homecookie = response.cookies
soup = BeautifulSoup(response.text, "lxml")
tokenTag=soup.find("input",{"name":"_token"})
token=tokenTag["value"]

#==============================================================================

loginurl = "https://companycheck.co.uk/login/login"
logindata = "email=sudheer.reddy@eclerx.com&password=eclerx#123&_token="+token
headers = {'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8','Content-Type':'application/x-www-form-urlencoded','Referer':'https://companycheck.co.uk/','Origin':'https://companycheck.co.uk','Host':'companycheck.co.uk','User-Agent':UserAgent}
response1 = requests.post(loginurl,data = logindata,headers = headers,cookies=Homecookie,proxies =proxy_support)
logincookie = response1.cookies
#print(logincookie)
# =============================================================================
# file=open("login.html","w+")
# file.write(str(response1.text.encode("utf-8")))
# file.close()
# #=============================================================================
# =============================================================================
#=============================================================================
wb = load_workbook(filename = 'E:\\Company\\Input\\15Jan.xlsx')
ws = wb['Client']
row = ws.max_row
col = ws.max_column
for row in ws.iter_rows(min_row=2, min_col=0, max_row=row, max_col=1):
    for cell in row:
        urlname.append(cell.value)
#=============================================================================
statusfile = "E:\\Company\\Input\\Status.xlsx"
workbook1 = xlsxwriter.Workbook(statusfile)
worksheet1 = workbook1.add_worksheet()
wr = 1
cl = 1
#==============================================================================
for cname in urlname:
    print("Search for"+cname)
    compheaders = {'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8','Referer':'https://companycheck.co.uk/dashboard','Host':'companycheck.co.uk','User-Agent':UserAgent}
    searchurl = "https://companycheck.co.uk/search?term="+cname
    searchresponse = requests.get(searchurl,cookies=logincookie, headers=compheaders ,proxies = proxy_support)
    searchCookie = searchresponse.cookies
    #print(searchresponse.status_code)
    if searchresponse.status_code != 200:
        continue
    soup = BeautifulSoup(searchresponse.text, "lxml")
    st = soup.find("div",{"class":"search-information"}).text
    if "0 results found" == st:
        worksheet1.write(wr, 0, cname)
        worksheet1.write(wr, 1, cname)
        worksheet1.write(wr, 2, "Not Found")
        wr+=1
        continue
    comdivMul = soup.find_all("div", {"itemprop": "itemListElement"})
    cnt = 0
    for comdiv in comdivMul:
#        cnt+=1
#        print(cnt)
        linkanchor = comdiv.find("a", {"class": "result__title"})
        link = linkanchor["href"]
        compname = linkanchor["title"]
#        if "result__status" in comdiv.text:
    #        if compname != cname:
    #            break
        statusTag = comdiv.find("span", {"class": "result__status"})
        if statusTag != []:
            status = statusTag.text
#            print(statusTag)          
            
            #==============================================================================
            compLink ="https://companycheck.co.uk"+link+"/financials"
            CompPage = requests.get(compLink,cookies =searchCookie, proxies = proxy_support)
            page = cname+".html"
#            print(CompPage.status_code)
            if CompPage.status_code != 200:
                worksheet1.write(wr, 0, cname)
                worksheet1.write(wr, 1, compname)
                worksheet1.write(wr, 2, "Not Found Search for next match")
                wr+=1
                continue
        # =============================================================================
        #     file=open(page,"w+")
        #     file.write(str(CompPage.text.encode("utf-8")))
        #     file.close()
        # =============================================================================
            #=====Code to create Excel file#===========
            Cpath = "E:\\Company\\15Jan_DataOut\\Client\\"
            if not os.path.exists(Cpath):
                    os.makedirs(Cpath)
            if "B/E" in cname:
                cname = cname.replace("/", ".")
            
            pagehtml = CompPage.text.replace("\n","")
            #==========================================
            soup = BeautifulSoup(CompPage.text, "lxml")
            tables = soup.find_all("table", {"class": "default-table"})
#            print(tables)
            if tables != []:
                filename = Cpath + compname + "-" + status + ".xlsx"
                print(filename)
                workbook = xlsxwriter.Workbook(filename)
                worksheet = workbook.add_worksheet()
                row = 1
                col = 0
                for table in tables:
                    for thead in table.find_all("thead"):
                        for td in thead.children:
                            data = td.string
#                            print(data)
                            if data != '\n':
                                worksheet.write(row, col, data)
                                col += 1
                        col = 0
                        row += 1
                    for tr in table.find_all("tr"):
                        for td in tr.children:
                            data = td.string
                            if data != '\n':
                                worksheet.write(row, col, data)
                                col += 1
                        row += 1
                        col = 0
                    col = 0
                    row = row+2
                workbook.close()
                worksheet1.write(wr, 0, cname)
                worksheet1.write(wr, 1, compname)
                worksheet1.write(wr, 2, "Data Found")
                wr += 1
                break
            else:
                worksheet1.write(wr, 0, cname)
                worksheet1.write(wr, 1, compname)
                worksheet1.write(wr, 2, "Table Not Found")
                wr += 1
                continue
# =============================================================================
#             else:
#                 worksheet1.write(wr, 0, cname)
#                 worksheet1.write(wr, 1, compname)
#                 worksheet1.write(wr, 2, " status Not Found")
#                 wr += 1
#                 continue
# =============================================================================
        else:
            worksheet1.write(wr, 0, cname)
            worksheet1.write(wr, 1, compname)
            worksheet1.write(wr, 2, " status Not Found")
            wr += 1
            continue