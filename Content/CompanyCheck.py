# -*- coding: utf-8 -*-
"""
Created on Wed Jan 10 12:20:47 2018

@author: Tech
"""
import requests
import os
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import xlsxwriter
import random

UserAgent = "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.94 Safari/537.36"
Company = "Global Payments U.K. Ltd"
compname = Company.replace(" ", "+")
urlname = []

#==============================================================================

proxy_support = {'https': 'https://csimonra:h19VA2xZ@173.246.181.89:29842',
                 'https': 'https://csimonra:h19VA2xZ@208.110.59.137:29842',
                 'https': 'https://csimonra:h19VA2xZ@208.110.59.208:29842',
                 'https': 'https://csimonra:h19VA2xZ@217.182.223.112:29842',
                 'https': 'https://csimonra:h19VA2xZ@217.182.223.118:29842',
                 'https': 'https://csimonra:h19VA2xZ@217.182.223.126:29842',
                 'https': 'https://csimonra:h19VA2xZ@173.228.169.148:29842'}
#==============================================================================
homeurl = "https://companycheck.co.uk/"
response = requests.get(homeurl, proxies=proxy_support)
Homecookie = response.cookies
soup = BeautifulSoup(response.text, "lxml")
tokenTag = soup.find("input", {"name": "_token"})
token = tokenTag["value"]

#==============================================================================

loginurl = "https://companycheck.co.uk/login/login"
logindata = "email=nitin.sawant@eclerx.com&password=eclerx#123&_token="+token
headers = {'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8', 'Content-Type': 'application/x-www-form-urlencoded', 'Referer': 'https://companycheck.co.uk/', 'Origin': 'https://companycheck.co.uk', 'Host': 'companycheck.co.uk', 'User-Agent': UserAgent}
response1 = requests.post(loginurl, data = logindata, headers = headers, cookies = Homecookie, proxies = proxy_support)
logincookie = response1.cookies

#=============================================================================

wb= load_workbook(filename = 'E:\\Company\\Input\\12Jan.xlsx')
ws = wb['Client']
row = ws.max_row
col=ws.max_column
for row in ws.iter_rows(min_row=2, min_col=0, max_row=row, max_col=1):
    for cell in row:
        urlname.append(cell.value)

#==============================================================================

for cname in urlname:
    compheaders = {'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8','Referer':'https://companycheck.co.uk/dashboard','Host':'companycheck.co.uk','User-Agent':UserAgent}
    searchurl ="https://companycheck.co.uk/search?term="+cname
    searchresponse = requests.get(searchurl,cookies=logincookie, headers=compheaders ,proxies = proxy_support)
    searchCookie = searchresponse.cookies
    #print(searchresponse.status_code)
    if searchresponse.status_code != 200:
        continue
    soup = BeautifulSoup(searchresponse.text, "lxml")
    st = soup.find("div",{"class":"search-information"}).text
#    print(st)
    if "0 results found" in st:
        continue
    comdivMul = soup.find_all("div",{"itemprop":"itemListElement"})
    for comdiv in comdivMul:
        linkanchor = comdiv.find("a",{"class":"result__title"})
        link = linkanchor["href"]
#        print(link)
        compname = linkanchor["title"]
#        if compname != cname:
#            break
        compSet = comdiv.find("div",{"class":"info-right"}).text
        status = comdiv.find("span",{"class":"result__status"}).text
        #==============================================================================
        compLink ="https://companycheck.co.uk"+link+"/financials"
        CompPage = requests.get(compLink,cookies =searchCookie, proxies = proxy_support)
        Cpath = "E:\\Company\\12Jan_DataOut\\Client\\"
        page = Cpath+cname+".html"
#            print(CompPage.status_code)
        if CompPage.status_code != 200:
            continue
        file=open(page,"w+")
        file.write(str(CompPage.text.encode("utf-8")))
        file.close()        
        #=====Code to create Excel file#===========
        
        if not os.path.exists(Cpath):
                os.makedirs(Cpath)
        if "B/E" in cname:
            cname=cname.replace("/",".")
        
#            pagehtml=CompPage.text.replace("\n","")
        #==========================================
        soup = BeautifulSoup(CompPage.text, "lxml")
        tables = soup.find_all("table",{"class":"default-table"})
        filename=Cpath+compname+"-"+status+".xlsx"
        print(filename)
        #file=open(filename,"w+")
        workbook = xlsxwriter.Workbook(filename)
        worksheet = workbook.add_worksheet()
        #worksheet.write('A1',Company,'bold')
        row = 1
        col = 0
        for table in tables:
#            print(table.text)
            for thead in table.find_all("thead"):
                for td in thead.children:
                    data = td.string
#                            print(data)
                    if data != '\n':
                        worksheet.write(row,col,data)
                        col+=1
                col=0
                row+=1
            for tr in table.find_all("tr"):
                for td in tr.children:
                    data = td.string
                    if data != '\n':
                        worksheet.write(row,col,data)
                        col+=1
                row+=1
                col=0
            col=0
            row=row+2
                
        workbook.close()
        break