from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
import xlsxwriter
catdata=[]

proxy_support ={'https': 'https://csimonra:h19VA2xZ@173.228.169.137:29842',
    'https': 'https://csimonra:h19VA2xZ@173.228.169.148:29842'}

wb= load_workbook(filename = 'E:\\Elevania\\count.xlsx')
ws = wb['Sheet1']
row = ws.max_row
col=ws.max_column
data=[]
#===========code to iterate through the excel file=============================
for row in ws.iter_rows(min_row=2, min_col=0, max_row=row, max_col=1):
    for cell in row:
        data.append(cell.value)#-----Store data to data list
#print(data)
file = open("E:\\URL.text",'w+')

for d in data:
    rs= requests.get(d,proxies=proxy_support)
    soup = BeautifulSoup(rs.text, "lxml")
    cnt = soup.find("b", {"class": "notranslate"}).text   
    cnt=cnt.replace(".","")
    pr = int(cnt)
    pages = int( pr/60)
    if pr%60 >0:
        pages+=1
    div= int(pages / 50)
    if pages%50:
        div+=1
    pno=1
    link=[]
    for i in range(div):
        link.append("ID|No:"+str(pno)+d+"|ASG_SCName|IDASG_SCID|ASG_Category|ASG_TID|NR|COMPETE")
        pno=pno+50
# =============================================================================
#     file.write(d)
#     file.write("-")
#     file.write(cnt)
#     file.write("-")
#     file.write(str(pages))
#     file.write("-")
#     file.write(str(div))
#     file.write("|")
# =============================================================================
    
    for l in link:
        file.write(l)
        file.write("}")
    print(d,cnt)
    
file.close()

