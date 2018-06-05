
import xlsxwriter
from openpyxl import load_workbook
dataList = []
Cat_1_Count = []
# Cat_2_URL = []
# Cat_2_Count = []
# Cat_3_URL = []
# Cat_3_Count = []

Cpath = "E:\\Hemant Python\\Qoo10\\"
fname = Cpath + "11streetMYurlsCount.xlsx"
wb = load_workbook(filename=fname)
ws = wb['Sheet2']
row = ws.max_row
col = ws.max_column
for row in ws.iter_rows(min_row=2, min_col=0, max_row=row, max_col=col):
    data = []
    for cell in row:
        data.append(cell.value)
        # print(data)
    dataList.append(data)
wb.close()


oname = Cpath + "Outfile.xlsx"
subURLBook = xlsxwriter.Workbook(oname)
subCatSplitSheet = subURLBook.add_worksheet(name='Split URL')
Cr = 1
subCatSplitSheet.write(Cr, 0, 'URL')
subCatSplitSheet.write(Cr, 1, 'Cat Count')
subCatSplitSheet.write(Cr, 2, 'Pages')
subCatSplitSheet.write(Cr, 3, 'Split URL Part')
subCatSplitSheet.write(Cr, 4, 'Split URL')
Cr += 1
part1 = "MY|No:"
part2 = "|asg|asg|Fashion|NA|NR|Seller"
for d in dataList:
    link = d[0]
    cnt = int(d[1])
    pno = 1
    pages = int(cnt/120)
    if cnt % 120 > 0:
        pages += 1
    PgSplitCnt = int(pages / 30)
    if pages % 30 > 0:
        PgSplitCnt += 1
    for j in range(0,PgSplitCnt):
        SplitURL = part1 + str(pno) + link + part2
        print(SplitURL)
        subCatSplitSheet.write(Cr, 0, link)
        subCatSplitSheet.write(Cr, 1, cnt)
        subCatSplitSheet.write(Cr, 2, pages)
        subCatSplitSheet.write(Cr, 3, PgSplitCnt)
        subCatSplitSheet.write(Cr, 4, SplitURL)
        pno += 30
        Cr += 1
# =============================================================================
subURLBook.close()
