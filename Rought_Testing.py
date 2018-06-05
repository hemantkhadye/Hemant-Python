# -*- coding: utf-8 -*-

# -*- coding: utf-8 -*-
"""
Created on Fri Jan 12 13:02:24 2018

@author: Tech
"""

import requests
import os
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import xlsxwriter

UserAgent = "Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.94 Safari/537.36"
Company = "Global Payments U.K. Ltd"
compname = Company.replace(" ", "+")
urlname = []
# =============================================================================
proxy_support = {'https':'https://csimonra:h19VA2xZ@173.246.181.89:29842',
                 'https':'https://csimonra:h19VA2xZ@208.110.59.137:29842',
                 'https':'https://csimonra:h19VA2xZ@208.110.59.208:29842',
                 'https':'https://csimonra:h19VA2xZ@217.182.223.112:29842',
                 'https':'https://csimonra:h19VA2xZ@217.182.223.118:29842',
                 'https':'https://csimonra:h19VA2xZ@217.182.223.126:29842',
                 'https':'https://csimonra:h19VA2xZ@173.228.169.148:29842'}
#==============================================================================
homeurl = "https://companycheck.co.uk/"
response = requests.get(homeurl,proxies = proxy_support)
Homecookie = response.cookies
soup = BeautifulSoup(response.text, "lxml")
tokenTag=soup.find("input",{"name":"_token"})
token=tokenTag["value"]

#==============================================================================

loginurl = "https://companycheck.co.uk/login/login"
logindata = "email=sudheer.reddy@eclerx.com&password=eclerx#123&_token="+token
headers = {'Accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8','Content-Type':'application/x-www-form-urlencoded','Referer':'https://companycheck.co.uk/','Origin':'https://companycheck.co.uk','Host':'companycheck.co.uk','User-Agent':UserAgent}
response1 = requests.post(loginurl,data = logindata,headers = headers,cookies=Homecookie,proxies =proxy_support)
logincookie = response1.cookies
#print(logincookie)
# =============================================================================
# file=open("login.html","w+")
# file.write(str(response1.text.encode("utf-8")))
# file.close()
# #=============================================================================
# =============================================================================
#=============================================================================
wb = load_workbook(filename = 'E:\\Company\\Input\\15Jan_URL.xlsx')
ws = wb['Client']
row = ws.max_row
col = ws.max_column
for row in ws.iter_rows(min_row=2, min_col=2, max_row=row, max_col=col):
    for cell in row:
        urlname.append(cell.value)
#print(urlname)
#==============================================================================
for url in urlname:
        CompPage = requests.get(url, cookies=logincookie, proxies = proxy_support)
        soup = BeautifulSoup(CompPage.text, "lxml")
        cname = soup.find("h1", {"class": "Company-header__name"}).text
        print(cname)
        page = "E:\\Company\\Input\\" + cname + ".html"
        file = open(page, "w+")
        file.write(str(CompPage.text.encode("utf-8")))
        file.close()
#        print(CompPage.status_code)
        if CompPage.status_code != 200:
            continue
        #=====Code to create Excel file#===========
        Cpath = "E:\\Company\\15Jan_DataOut\\Client\\"
        if not os.path.exists(Cpath):
                os.makedirs(Cpath)
        if "B/E" in cname:
            cname = cname.replace("/", ".")
        
        pagehtml = CompPage.text.replace("\n","")
        #==========================================
        
        tables = soup.find_all("table", {"class": "default-table"})
        print(tables)
        if tables != []:
            filename = Cpath +cname + ".xlsx"
            print(filename)
            workbook = xlsxwriter.Workbook(filename)
            worksheet = workbook.add_worksheet()
            row = 1
            col = 0
            for table in tables:
                print(table.text)
                for thead in table.find_all("thead"):
                    for td in thead.children:
                        data = td.string
#                            print(data)
                        if data != '\n':
                            worksheet.write(row, col, data)
                            col += 1
                    col = 0
                    row += 1
                for tr in table.find_all("tr"):
                    for td in tr.children:
                        data = td.string
                        if data != '\n':
                            worksheet.write(row, col, data)
                            col += 1
                    row += 1
                    col = 0
                col = 0
                row = row+2
            workbook.close()
           
